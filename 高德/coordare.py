# -*- coding: utf-8 -*-

import json

import tablib as tablib
from openpyxl import Workbook
from pandas.io.json import json_normalize
def json_to_excel(json_file,excel_file):
    #读取json文件数据
    with open(json_file, mode='r', encoding='utf-8') as jf:
        json_data= json.load(jf)
    k = json_data.keys()
    wb = Workbook()#创建excel文件
    for sheet_name in k:
        try:
            wb.remove(sheet_name)  # 如表已存在则移除工作表
        except:
            pass
        wb.create_sheet(sheet_name, 0)  # 创建表
        ws = wb[sheet_name]  # 操作指定表
        sheet_data = json_data[sheet_name]  # 获取表要写入的数据
        for t in range(1, len(sheet_data) + 1):  # 遍历要写入的行数
            i = 65  # ASCII码'65'表示'A'
            for j in range(1, len(sheet_data[t - 1]) + 1):  # 遍历每行要写入的数量
                ws['%s%d' % (chr(i), t)] = sheet_data[t - 1][j - 1]
                i += 1
    wb.save(excel_file)



file_name = 'coor_res.txt'
with open(file_name,'r',encoding='utf-8') as f:
    rows = f.readlines()
    coor_list =[]
    for row in rows:
        coor_json = json.loads(row)
        poi = coor_json["pois"]
        poiId = coor_json["poiId"]
        poiName = coor_json["poiName"]
        #city = coor_json["city"]
        if len(poi)>0:
            address = poi[0]["address"]
            name = poi[0]["name"]
            location = poi[0]["location"].split(',')
            lon = location[0]
            lat = location[1]
            w_json = {
                "status:":1,
                "address":address,
                "name":name,
                "location":location,
                "lon":lon,
                "lat":lat,
                "poiId":poiId,
                "poiName":poiName
                #"city":city
            }
        else:
            w_json={
                "status":0,
                "address": "",
                "name": "",
                "location": "",
                "lon": "",
                "lat": "",
                "poiId": poiId,
                "poiName": poiName
                #"city":city
            }
        coor_list.append(w_json)
with open('result/coor_excle.txt', 'w', encoding='utf-8') as a:
     for row in coor_list:
         #txt = json.dumps(row, ensure_ascii=False)
         txt = str(row["address"])+'\t'+str(row["name"])+'\t'+str(row["lon"])+'\t'+str(row["lat"])+'\t'+str(row["poiId"])+'\t'+str(row["poiName"])+'\n'
         a.write(txt)
    # 格式化
    # dat_table = json_normalize(row)
    #
    # # 导出excel
    # dat_table.to_excel('./coor_result.xlsx','a+', index=False)

# header=tuple([ i for i in coor_list[0].keys()])
# data = []
# # 循环里面的字典，将value作为数据写入进去
# for row in rows:
#      body = []
#      for v in row.values():
#          body.append(v)
#          data.append(tuple(body))
# #将含标题和内容的数据放到data里
# data = tablib.Dataset(*data,headers=header)
# #写到桌面
# open('coor_res', 'wb').write(data.xls)